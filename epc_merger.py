# EPC Merger and Reconciliation Tool (Index-Based Column Selection - Batch Smart)

import pandas as pd
import os
import csv
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox

all_batches = []

def select_files():
    root = tk.Tk()
    root.withdraw()
    file_paths = filedialog.askopenfilenames(
        title="Select a Batch of RFID Scan Files",
        filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")]
    )
    return list(file_paths)

def choose_epc_column_gui(columns, filename, preview=None):
    root = tk.Tk()
    root.title("Select EPC Column")
    msg = f"{filename}\nSelect the column that contains EPCs:"
    if preview:
        msg += f"\n\nPreview:\n{preview}"
    tk.Label(root, text=msg).pack(padx=10, pady=10)

    var = tk.StringVar(root)
    var.set(columns[0])
    dropdown = tk.OptionMenu(root, var, *columns)
    dropdown.pack(padx=10, pady=10)

    def on_submit():
        root.quit()
        root.destroy()

    tk.Button(root, text="Confirm", command=on_submit).pack(pady=10)
    root.mainloop()

    return var.get()

def is_epc_like(value):
    if isinstance(value, str):
        value = value.strip().replace(' ', '')
        return value.isalnum() and len(value) >= 8
    return False

def read_file_flexible(file, nrows=None):
    try:
        # Try to detect a header row
        with open(file, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        header_index = None
        for i, line in enumerate(lines):
            lowered = line.lower()
            if ("epc" in lowered or "tag" in lowered) and "count" in lowered:
                header_index = i
                break

        if header_index is not None:
            df = pd.read_csv(file, delimiter=',', skiprows=header_index, on_bad_lines='skip', nrows=nrows)
        else:
            print(f"⚠️ No header found in {file}, treating as raw EPC list.")
            df = pd.read_csv(file, delimiter=',', header=None, names=["EPC"], on_bad_lines='skip', nrows=nrows)

        return df

    except Exception as e:
        print(f"❌ Failed to read file {file}: {e}")
        return pd.DataFrame()




def load_batch(files):
    preview_df = read_file_flexible(files[0], nrows=10)
    preview_df = preview_df.dropna(axis=1, how='all').dropna(axis=0, how='all')
    preview_rows = preview_df.head(3).astype(str).agg(', '.join, axis=1).tolist()
    preview_text = "\n".join(preview_rows)

    columns_list = preview_df.columns.tolist()
    if not columns_list:
        print(f"❌ No usable columns found in: {files[0]}")
        return  # Skip this batch

    epc_col = choose_epc_column_gui([str(col) for col in columns_list], Path(files[0]).name, preview_text)

    epc_index = columns_list.index(epc_col)

    root = tk.Tk()
    root.withdraw() 
    prefix_filters = []
    while True:
        prefix = simpledialog.askstring("Filter EPCs", "Enter EPC prefix(es) to include (e.g. 01, 03).\nPress Enter without typing to finish.")
        if prefix:
            parts = [p.strip() for p in prefix.split(",") if p.strip()]
            prefix_filters.extend(parts)
        else:
            break

    # Ask for EPC character limit (after prefix loop finishes)
    char_limit = simpledialog.askinteger(
        "Truncate EPCs",
        "Enter number of characters to keep from each EPC (e.g. 24).\nLeave blank to use full EPCs:"
    )




    batch_epcs = []
    for file in files:
        print(f"Reading file: {file}")
        df = read_file_flexible(file)
        df = df.dropna(axis=1, how='all').dropna(axis=0, how='all')

        if epc_index >= len(df.columns):
            print(f"❌ Skipping {file} — column index {epc_index} out of range.")
            continue

        selected_col = df.columns[epc_index]
        df = df[[selected_col]].dropna()
        df.columns = ["EPC"]

        # Truncate EPCs if a limit is set
        if char_limit:
            df["EPC"] = df["EPC"].astype(str).str[:char_limit]


        df = df[df["EPC"].apply(is_epc_like)]

        if prefix_filters:
            df = df[df["EPC"].str.startswith(tuple(prefix_filters))]


        # Extract metadata from filename
        parts = Path(file).stem.split("_")
        df["Reader"] = parts[0] if len(parts) > 0 else "Unknown"
        df["Location"] = parts[1] if len(parts) > 1 else "Unknown"
        df["File Name"] = Path(file).stem

        batch_epcs.append(df)

    if batch_epcs:
        merged_batch = pd.concat(batch_epcs, ignore_index=True)
        all_batches.append(merged_batch)
        print(f"✅ Batch of {len(files)} files added.")

                # ✅ Automatically mark folder as merged
        folder_path = os.path.dirname(files[0])
        parent = Path(folder_path).parent
        current = Path(folder_path).name

        if "✔" not in current:
            new_name = current + " ✔"
            try:
                os.rename(folder_path, str(parent / new_name))
                print(f"📁 Renamed folder to: {new_name}")
            except Exception as e:
                print(f"⚠️ Could not rename folder: {e}")


    else:
        print("⚠️ No valid EPC data found in this batch.")

if __name__ == "__main__":
    print("Starting EPC Merger Tool (Batch Mode)...")
    while True:
        selected_files = select_files()
        if not selected_files:
            break
        load_batch(selected_files)
        add_more = messagebox.askyesno("Add Another Batch?", "Do you want to load another batch of files?")
        if not add_more:
            break

    if all_batches:
        confirm_merge = messagebox.askyesno("Confirm Merge", "Are you sure you want to merge all uploaded batches and save?")
        if confirm_merge:
            final_merged = pd.concat(all_batches, ignore_index=True)
            final_merged = final_merged.drop_duplicates(subset="EPC").sort_values("EPC")
            from datetime import datetime
            os.makedirs("merged", exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"merged/Merged_Cleaned_EPCs_{timestamp}.xlsx"
            final_merged.to_excel(filename, index=False)

            # Auto-adjust Excel column widths
            import openpyxl
            from openpyxl.utils import get_column_letter

            try:
                wb = openpyxl.load_workbook(filename)
                ws = wb.active

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column  # 1-based index
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[get_column_letter(column)].width = max_length + 2

                wb.save(filename)
                print(f"✅ All batches merged and saved to '{filename}'")

            except PermissionError:
                print(f"❌ Cannot save. Please close the file '{filename}' and try again.")

        else:   
            print("❌ Merge cancelled. No file was saved.")
    else:
        print("❌ No EPC data was merged.")
