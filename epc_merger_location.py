import pandas as pd
import os
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
import threading
import subprocess

all_batches = []

def select_files_or_folder():
    root = tk.Tk()
    root.attributes("-topmost", True)
    root.withdraw()
    choice = messagebox.askyesno("Select Folder?", "Do you want to select a folder?\n(Yes = folder, No = individual files)")
    if choice:
        folder_path = filedialog.askdirectory(title="Select Folder to Search Files In")
        if folder_path:
            file_paths = []
            for dirpath, _, filenames in os.walk(folder_path):
                for filename in filenames:
                    if filename.endswith((".xlsx", ".xls", ".csv")):
                        file_paths.append(os.path.join(dirpath, filename))
            return file_paths
        else:
            return []
    else:
        file_paths = filedialog.askopenfilenames(
            title="Select a Batch of RFID Scan Files",
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")]
        )
        return list(file_paths)

def read_file_flexible(file, nrows=None):
    try:
        # Read raw lines
        with open(file, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        def is_epc_like(val):
            val = val.strip()
            return val.isalnum() and len(val) >= 16 and not val.isdigit()

        # Find the row index where actual EPC-like data starts
        data_start_row = None
        for i, line in enumerate(lines):
            cells = line.strip().split(',')
            if any(is_epc_like(cell) for cell in cells):
                data_start_row = i
                break

        if data_start_row is None:
            print("❌ No EPC-like data found in file.")
            return pd.DataFrame()

        # Read from detected start row (skip preceding summary rows)
        df = pd.read_csv(file, header=None, skiprows=data_start_row, on_bad_lines='skip', nrows=nrows)
        return df

    except Exception as e:
        print(f"❌ Failed to read file {file}: {e}")
        return pd.DataFrame()


def choose_epc_column_gui_with_preview(df, filename):
    from tkinter import ttk
    import re

    def is_epc_like(val):
        val = str(val).strip()
        return val.isalnum() and len(val) >= 16 and not val.isdigit()

    root = tk.Tk()
    root.attributes("-topmost", True)
    root.title("Select EPC Column")

    tk.Label(root, text=f"{filename}\nSelect the column that contains EPCs:").pack(padx=10, pady=10)

    column_choices = []
    col_index_map = {}
    epc_preview = {}
    epc_scores = {}

    # Ensure all columns up to at least index 19 exist
    for col in range(20):
        if col not in df.columns:
            df[col] = None

    for idx in sorted(df.columns):
        series = df[idx].dropna().astype(str).str.strip()
        series_list = series.tolist()

        # Detect EPC-like values
        epc_vals = [v for v in series_list if is_epc_like(v)]
        epc_scores[idx] = len(epc_vals)

        # Show both EPC and fallback values
        fallback_vals = series_list[:4] if series_list else ["Empty"]
        preview_vals = epc_vals[:4] if epc_vals else fallback_vals[:4]

        # Use the same 1st to 3rd values for both dropdown and preview
        epc_preview[idx] = preview_vals[:3]  # Top preview
        label = f"[{idx}] {', '.join(preview_vals[:3])[:40]}"  # Dropdown

        column_choices.append(label)
        col_index_map[label] = idx

    # Handle case where no EPCs are found at all
    if not epc_scores:
        messagebox.showerror("No EPC Found", "No EPC-like values detected in any column.")
        root.destroy()
        return 0

    best_col = max(epc_scores, key=epc_scores.get)

    # Highlight the best one
    updated_choices = []
    auto_detected_label = None
    for label in column_choices:
        idx = col_index_map[label]
        if idx == best_col:
            label = f"✔️ {label}"
            auto_detected_label = label
        updated_choices.append(label)

    var = tk.StringVar(root)
    var.set(auto_detected_label or updated_choices[0])

    preview_label = tk.Label(root, text="", justify="left", fg="blue", font=("Consolas", 10))
    preview_label.pack(padx=10, pady=5)

    combobox = ttk.Combobox(root, textvariable=var, values=updated_choices, state="readonly", width=60)
    combobox.pack(padx=10, pady=10)

    def update_preview(*args):
        selected_label = var.get().replace("✔️ ", "")
        selected_index = col_index_map[selected_label]
        preview = "\n".join(epc_preview[selected_index])
        preview_label.config(text=f"📄 Preview from column {selected_index}:\n{preview}")

    var.trace_add('write', update_preview)
    update_preview()

    def on_submit():
        root.quit()
        root.destroy()

    tk.Button(root, text="Confirm", command=on_submit).pack(pady=10)
    root.mainloop()

    selected_label = var.get().replace("✔️ ", "")
    return col_index_map[selected_label]



def show_loading_popup():
    popup = tk.Toplevel()
    popup.title("Merging...")
    popup.geometry("320x100")
    popup.configure(bg="#F4F4F4")
    popup.attributes("-topmost", True)
    popup.grab_set()

    tk.Label(
        popup,
        text="Processing and saving merged EPCs...",
        font=("Segoe UI", 11),
        bg="#F4F4F4"
    ).pack(expand=True, pady=30)

    popup.update()
    return popup

def open_merged_folder():
    folder = os.path.abspath("merged")
    try:
        subprocess.Popen(["cmd", "/c", "start", "", folder], shell=True)
    except Exception as e:
        print(f"⚠️ Could not open folder: {e}")

def save_and_exit():
    if not all_batches:
        print("❌ No data merged.")
        return

    popup = show_loading_popup()
    done_event = threading.Event()

    def merge_and_save():
        from datetime import datetime
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter

            merged_all = pd.concat(all_batches, ignore_index=True)

            # Group by EPC and combine all unique locations for each EPC
            merged_grouped = (
                merged_all.groupby("EPC")
                .agg({
                    "Location": lambda x: ", ".join(sorted(set(x))),
                    "File Name": lambda x: ", ".join(sorted(set(x)))  # Optional: also merge file names
                })
                .reset_index()
                .sort_values("EPC")
            )
            final_merged = merged_grouped

            os.makedirs("merged", exist_ok=True)
            filename = f"merged/Merged_EPCs_LocationOnly_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            final_merged.to_excel(filename, index=False)

            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            for col in ws.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
            wb.save(filename)
            print(f"✅ File saved to: {filename}")
        except Exception as e:
            print(f"❌ Merge failed: {e}")
        done_event.set()

    def check_if_done():
        if done_event.is_set():
            popup.destroy()
            open_merged_folder()
        else:
            popup.after(100, check_if_done)

    threading.Thread(target=merge_and_save).start()
    check_if_done()

def load_batch(files):
    preview_df = read_file_flexible(files[0], nrows=50)
    preview_df = preview_df.dropna(axis=1, how='all').dropna(axis=0, how='all')
    if preview_df.empty or len(preview_df.columns) == 0:
        print(f"❌ No usable columns found in: {files[0]}")
        return

    epc_index = choose_epc_column_gui_with_preview(preview_df, Path(files[0]).name)

    root = tk.Tk()
    root.attributes("-topmost", True)
    root.withdraw()
    prefix_filters = []
    while True:
        prefix = simpledialog.askstring("Filter EPCs", "Enter EPC prefix(es) to include (e.g. 01, 03). Leave blank to skip.")
        if prefix:
            parts = [p.strip() for p in prefix.split(",") if p.strip()]
            prefix_filters.extend(parts)
        else:
            break

    char_input = simpledialog.askstring("Truncate EPCs", "Enter number of characters to keep (e.g. 24), or leave blank:")
    char_limit = int(char_input) if char_input and char_input.isdigit() else None

    batch_epcs = []
    for file in files:
        df = read_file_flexible(file)
        df = df.dropna(axis=1, how='all').dropna(axis=0, how='all')
        if epc_index >= len(df.columns):
            print(f"❌ Skipping {file} — column index {epc_index} out of range.")
            continue

        selected_col = df.columns[epc_index]
        df = df[[selected_col]].dropna()
        df.columns = ["EPC"]

        if char_limit:
            df["EPC"] = df["EPC"].astype(str).str[:char_limit]
        if prefix_filters:
            df = df[df["EPC"].str.startswith(tuple(prefix_filters))]

        file_stem = Path(file).stem
        df["Location"] = file_stem.split("_", 1)[0]
        df["File Name"] = file_stem
        batch_epcs.append(df)

    if batch_epcs:
        merged_batch = pd.concat(batch_epcs, ignore_index=True)
        all_batches.append(merged_batch)
        print(f"✅ Added {len(files)} files.")
    else:
        print("⚠️ No valid data found.")

if __name__ == "__main__":
    print("📦 EPC Merger (Location + File Name with EPC detection + loading popup)")
    while True:
        files = select_files_or_folder()
        if not files:
            print("❌ No files selected.")
            break
        load_batch(files)
        if not messagebox.askyesno("More?", "Load another batch?"):
            break

    if all_batches:
        if messagebox.askyesno("Confirm Merge", "Merge all batches and save file?"):
            save_and_exit()
