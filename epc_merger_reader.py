# epc_merger_reader.py (Updated with Smart EPC Detection and Preview)

import pandas as pd
import os
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from tkinter import ttk
import threading

all_batches = []

def select_files_or_folder():
    root = tk.Tk()
    root.attributes("-topmost", True)
    root.withdraw()
    choice = messagebox.askyesno("Select Folder?", "Do you want to select a folder?\n(Yes = folder, No = individual files)")
    if choice:
        folder_path = filedialog.askdirectory(title="Select Folder to Search Files In")
        if folder_path:
            file_paths = []
            for dirpath, _, filenames in os.walk(folder_path):
                for filename in filenames:
                    if filename.endswith((".xlsx", ".xls", ".csv")):
                        file_paths.append(os.path.join(dirpath, filename))
            return file_paths
        else:
            return []
    else:
        file_paths = filedialog.askopenfilenames(
            title="Select a Batch of RFID Scan Files",
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")]
        )
        return list(file_paths)

def read_file_flexible(file, nrows=None):
    try:
        with open(file, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        def is_epc_like(val):
            val = val.strip()
            return val.isalnum() and len(val) >= 16 and not val.isdigit()

        data_start_row = None
        for i, line in enumerate(lines):
            cells = line.strip().split(',')
            if any(is_epc_like(cell) for cell in cells):
                data_start_row = i
                break

        if data_start_row is None:
            print(f"❌ No EPC-like data found in {file}")
            return pd.DataFrame()

        df = pd.read_csv(file, header=None, skiprows=data_start_row, on_bad_lines='skip', nrows=nrows)
        return df
    except Exception as e:
        print(f"❌ Failed to read file {file}: {e}")
        return pd.DataFrame()

def choose_epc_column_gui_with_preview(df, filename):
    def is_epc_like(val):
        val = str(val).strip()
        return val.isalnum() and len(val) >= 16 and not val.isdigit()

    root = tk.Tk()
    root.attributes("-topmost", True)
    root.title("Select EPC Column")

    tk.Label(root, text=f"{filename}\nSelect the column that contains EPCs:").pack(padx=10, pady=10)

    column_choices = []
    col_index_map = {}
    epc_preview = {}
    epc_scores = {}

    for col in range(20):
        if col not in df.columns:
            df[col] = None

    for idx in sorted(df.columns):
        series = df[idx].dropna().astype(str).str.strip()
        series_list = series.tolist()

        epc_vals = [v for v in series_list if is_epc_like(v)]
        epc_scores[idx] = len(epc_vals)

        fallback_vals = series_list[:4] if series_list else ["Empty"]
        preview_vals = epc_vals[:4] if epc_vals else fallback_vals[:4]

        epc_preview[idx] = preview_vals[:3]
        label = f"[{idx}] {', '.join(preview_vals[:3])[:40]}"
        column_choices.append(label)
        col_index_map[label] = idx

    if not epc_scores:
        messagebox.showerror("No EPC Found", "No EPC-like values detected in any column.")
        root.destroy()
        return 0

    best_col = max(epc_scores, key=epc_scores.get)
    updated_choices = []
    auto_detected_label = None
    for label in column_choices:
        idx = col_index_map[label]
        if idx == best_col:
            label = f"✔️ {label}"
            auto_detected_label = label
        updated_choices.append(label)

    var = tk.StringVar(root)
    var.set(auto_detected_label or updated_choices[0])

    preview_label = tk.Label(root, text="", justify="left", fg="blue", font=("Consolas", 10))
    preview_label.pack(padx=10, pady=5)

    combobox = ttk.Combobox(root, textvariable=var, values=updated_choices, state="readonly", width=60)
    combobox.pack(padx=10, pady=10)

    def update_preview(*args):
        selected_label = var.get().replace("✔️ ", "")
        selected_index = col_index_map[selected_label]
        preview = "\n".join(epc_preview[selected_index])
        preview_label.config(text=f"📄 Preview from column {selected_index}:\n{preview}")

    var.trace_add('write', update_preview)
    update_preview()

    def on_submit():
        root.quit()
        root.destroy()

    tk.Button(root, text="Confirm", command=on_submit).pack(pady=10)
    root.mainloop()

    selected_label = var.get().replace("✔️ ", "")
    return col_index_map[selected_label]

def load_batch(files):
    preview_df = read_file_flexible(files[0], nrows=50)
    preview_df = preview_df.dropna(axis=1, how='all').dropna(axis=0, how='all')
    if preview_df.empty or len(preview_df.columns) == 0:
        print(f"❌ No usable columns found in: {files[0]}")
        return

    epc_index = choose_epc_column_gui_with_preview(preview_df, Path(files[0]).name)

    root = tk.Tk()
    root.attributes("-topmost", True)
    root.withdraw()
    prefix_filters = []
    while True:
        prefix = simpledialog.askstring("Filter EPCs", "Enter EPC prefix(es) to include (e.g. 01, 03). Leave blank to skip.")
        if prefix:
            parts = [p.strip() for p in prefix.split(",") if p.strip()]
            prefix_filters.extend(parts)
        else:
            break

    char_input = simpledialog.askstring("Truncate EPCs", "Enter number of characters to keep (e.g. 24), or leave blank:")
    char_limit = int(char_input) if char_input and char_input.isdigit() else None

    batch_epcs = []
    skipped = []

    for file in files:
        df = read_file_flexible(file)
        df = df.dropna(axis=1, how='all').dropna(axis=0, how='all')

        if epc_index >= len(df.columns):
            print(f"❌ Skipped (column index out of range): {file}")
            skipped.append(file)
            continue

        selected_col = df.columns[epc_index]
        df = df[[selected_col]].dropna()
        df.columns = ["EPC"]

        if df.empty:
            print(f"⚠️ Skipped (empty EPC column): {file}")
            skipped.append(file)
            continue

        if char_limit:
            df["EPC"] = df["EPC"].astype(str).str[:char_limit]
        if prefix_filters:
            df = df[df["EPC"].str.startswith(tuple(prefix_filters))]

        file_stem = Path(file).stem
        segments = file_stem.split("_")
        reader = segments[0] if len(segments) >= 1 else "Unknown"
        location = segments[1] if len(segments) >= 2 else "Unknown"

        df["Reader"] = reader
        df["Location"] = location
        df["File Name"] = file_stem
        batch_epcs.append(df)

    if batch_epcs:
        merged_batch = pd.concat(batch_epcs, ignore_index=True)
        all_batches.append(merged_batch)
        print(f"✅ Merged {len(batch_epcs)} files.")
    else:
        print("⚠️ No valid data found.")

    if skipped:
        print(f"⚠️ Skipped {len(skipped)} files due to format issues:")
        for s in skipped:
            print(f" - {s}")

def show_loading_popup():
    popup = tk.Toplevel()
    popup.title("Merging...")
    popup.geometry("320x100")
    popup.configure(bg="#F4F4F4")
    popup.attributes("-topmost", True)
    popup.grab_set()

    tk.Label(
        popup,
        text="Processing and saving merged EPCs...",
        font=("Segoe UI", 11),
        bg="#F4F4F4"
    ).pack(expand=True, pady=30)

    popup.update()
    return popup

def save_and_exit():
    if not all_batches:
        print("❌ No data merged.")
        return

    popup = show_loading_popup()
    done_event = threading.Event()

    def merge_and_save():
        from datetime import datetime
        final_merged = pd.concat(all_batches, ignore_index=True).drop_duplicates(subset="EPC").sort_values("EPC")
        os.makedirs("merged", exist_ok=True)
        filename = f"merged/Merged_EPCs_Reader_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        final_merged.to_excel(filename, index=False)

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            for col in ws.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
            wb.save(filename)
            print(f"✅ File saved to: {filename}")
        except Exception as e:
            print(f"⚠️ Saved, but column auto-fit failed: {e}")

        done_event.set()

    threading.Thread(target=merge_and_save).start()

    def check_if_done():
        if done_event.is_set():
            popup.destroy()
            os.startfile(os.path.abspath("merged"))
        else:
            popup.after(100, check_if_done)

    check_if_done()

if __name__ == "__main__":
    print("📦 EPC Merger (Reader + Location mode, Smart EPC Detection)")
    while True:
        files = select_files_or_folder()
        if not files: break
        load_batch(files)
        if not messagebox.askyesno("More?", "Load another batch?"): break
    if all_batches:
        if messagebox.askyesno("Confirm Merge", "Merge all batches and save file?"):
            save_and_exit()
