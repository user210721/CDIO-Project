import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, simpledialog
from datetime import datetime
import threading
import subprocess

def select_excel_files():
    root = tk.Tk()
    root.attributes("-topmost", True)
    root.withdraw()
    file_paths = filedialog.askopenfilenames(
        title="Select Merged EPC Excel Files",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    return list(file_paths)

def show_loading_popup():
    popup = tk.Toplevel()
    popup.title("Merging...")
    popup.geometry("320x100")
    popup.configure(bg="#F4F4F4")
    popup.attributes("-topmost", True)
    popup.grab_set()
    tk.Label(
        popup,
        text="Processing and saving final merged EPCs...",
        font=("Segoe UI", 11),
        bg="#F4F4F4"
    ).pack(expand=True, pady=30)
    popup.update()
    return popup

def open_merged_folder():
    folder = os.path.abspath("merged_final")
    try:
        subprocess.Popen(["cmd", "/c", "start", "", folder], shell=True)
    except Exception as e:
        print(f"⚠️ Could not open folder: {e}")

def merge_cleaned_files(files):
    root = tk.Tk()
    root.attributes("-topmost", True)
    root.withdraw()

    # Prefix filter
    prefix_filters = []
    while True:
        prefix = simpledialog.askstring("Filter EPCs", "Enter EPC prefix(es) to include (e.g. 01, 03).\nPress Enter without typing to finish.")
        if prefix:
            parts = [p.strip() for p in prefix.split(",") if p.strip()]
            prefix_filters.extend(parts)
        else:
            break

    # Truncate EPC
    char_input = simpledialog.askstring("Truncate EPCs", "Enter number of characters to keep from each EPC (e.g. 24).\nLeave blank to use full EPCs:")
    try:
        char_limit = int(char_input)
    except (TypeError, ValueError):
        char_limit = None

    # Save name
    os.makedirs("merged_final", exist_ok=True)
    default_name = f"Final_Merged_EPCs_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    output_name = simpledialog.askstring("Save As", "Enter name for final merged file:", initialvalue=default_name)
    if not output_name:
        print("❌ Save cancelled.")
        return

    popup = show_loading_popup()
    done_event = threading.Event()

    def merge_and_save():
        try:
            dfs = []
            all_columns = set()

            # Step 1: Collect all column names (except "File Name")
            for file in files:
                try:
                    df = pd.read_excel(file)
                    if "EPC" not in df.columns:
                        print(f"❌ Skipping {file} — no EPC column found.")
                        continue
                    all_columns.update(col for col in df.columns if col != "File Name")
                except Exception as e:
                    print(f"❌ Error reading {file}: {e}")

            # Final column order
            all_columns = list(all_columns)
            all_columns = [col for col in all_columns if col != "EPC"]
            all_columns = ["EPC"] + sorted(all_columns, key=str.lower)

            # Step 2: Normalize and process each file
            for file in files:
                try:
                    df = pd.read_excel(file)

                    if "EPC" not in df.columns:
                        continue

                    df["EPC"] = df["EPC"].astype(str)

                    if char_limit:
                        df["EPC"] = df["EPC"].apply(lambda x: x[:char_limit] if len(x) > char_limit else x)

                    if prefix_filters:
                        df = df[df["EPC"].str.startswith(tuple(prefix_filters))]

                    for col in all_columns:
                        if col not in df.columns:
                            df[col] = "Unknown"
                        else:
                            df[col] = df[col].fillna("Unknown")

                    df = df[all_columns]
                    dfs.append(df)

                    print(f"✅ Loaded: {file} ({len(df)} rows after filtering)")

                except Exception as e:
                    print(f"❌ Error processing {file}: {e}")

            if not dfs:
                print("❌ No valid files to merge.")
                done_event.set()
                return

            combined = pd.concat(dfs, ignore_index=True)

            # Merge duplicates by EPC
            def merge_column(series):
                values = sorted(set(
                    str(i).strip() for i in series.dropna()
                    if str(i).strip().lower() != "unknown" and str(i).strip() != ""
                ))
                return ", ".join(values) if values else "Unknown"

            agg_dict = {col: merge_column for col in all_columns if col != "EPC"}
            merged = combined.groupby("EPC", as_index=False).agg(agg_dict)

            # Save to Excel
            save_path = os.path.join("merged_final", f"{output_name}.xlsx")
            merged.to_excel(save_path, index=False)

            # Auto-adjust column widths
            import openpyxl
            from openpyxl.utils import get_column_letter
            wb = openpyxl.load_workbook(save_path)
            ws = wb.active
            for col in ws.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
            wb.save(save_path)

            print(f"✅ Final merged file saved to: {save_path}")

        finally:
            done_event.set()

    def check_if_done():
        if done_event.is_set():
            popup.destroy()
            open_merged_folder()
        else:
            popup.after(100, check_if_done)

    threading.Thread(target=merge_and_save).start()
    check_if_done()

if __name__ == "__main__":
    selected_files = select_excel_files()
    if selected_files:
        merge_cleaned_files(selected_files)
    else:
        print("❌ No files selected.")
