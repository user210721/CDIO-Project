# EPC Master Comparison Tool v9
# Adds "Location Found" and "Reader Used" columns to outputs

import pandas as pd
import os
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from datetime import datetime

EXCEL_MAX_ROWS = 1048576  # Excel row limit

def select_files_and_folders(title):
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)  # ✅ Always on top

    choice = messagebox.askyesno(
        "Select Folder?",
        f"Do you want to select a folder for {title}?\n(Yes = folder, No = individual files)"
    )

    files = []
    if choice:
        folder_path = filedialog.askdirectory(title=f"Select Folder for {title}")
        if folder_path:
            for dirpath, _, filenames in os.walk(folder_path):
                for filename in filenames:
                    if filename.lower().endswith((".xlsx", ".xls", ".csv")):
                        files.append(os.path.join(dirpath, filename))
            if not files:
                messagebox.showerror("No Files Found", f"No Excel/CSV files found in {folder_path}")
                return []
    else:
        file_paths = filedialog.askopenfilenames(
            title=f"Select {title}",
            filetypes=[("Excel/CSV Files", "*.xlsx *.xls *.csv")]
        )
        files.extend(file_paths)

    return list(files)

def load_merged_epcs(files):
    """Load EPCs from merged files with Location and Reader info"""
    merged_data = {}
    for file in files:
        try:
            if file.lower().endswith(".csv"):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)

            if "EPC" not in df.columns:
                print(f"⚠️ Skipping {file} — no EPC column found.")
                continue

            df["EPC"] = df["EPC"].astype(str).str.strip()
            for _, row in df.iterrows():
                epc = row["EPC"]
                location = row.get("Location", "")
                reader = row.get("Reader", "")
                merged_data[epc] = {"Location": location, "Reader": reader}

        except Exception as e:
            print(f"❌ Error reading {file}: {e}")
    return merged_data

def auto_adjust_columns(file_path):
    """Auto-adjust column widths for Excel file"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        for col in ws.columns:
            max_length = 0
            column = col[0].column  # 1-based index
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[get_column_letter(column)].width = max_length + 2

        wb.save(file_path)
        print(f"📏 Auto-adjusted columns: {file_path}")

    except Exception as e:
        print(f"⚠️ Could not auto-adjust columns for {file_path}: {e}")

def process_master_file(master_file, merged_epcs, output_folder, summary_rows, totals):
    """Compare a single master file and save result"""
    try:
        if master_file.lower().endswith(".csv"):
            df = pd.read_csv(master_file)
        else:
            df = pd.read_excel(master_file)

        if "EPC" not in df.columns:
            print(f"⚠️ Skipping {master_file} — no EPC column found.")
            return

        df["EPC"] = df["EPC"].astype(str).str.strip()
        df["Found"] = df["EPC"].apply(lambda x: "✅" if x in merged_epcs else "")
        df["Location Found"] = df["EPC"].apply(lambda x: merged_epcs[x]["Location"] if x in merged_epcs else "")
        df["Reader Used"] = df["EPC"].apply(lambda x: merged_epcs[x]["Reader"] if x in merged_epcs else "")

        found_count = df["Found"].eq("✅").sum()
        total_rows = len(df)
        percent_found = (found_count / total_rows) * 100 if total_rows else 0

        # Update totals
        totals["total_rows"] += total_rows
        totals["found_rows"] += found_count

        # Save result file
        base_name = Path(master_file).stem
        output_file = os.path.join(output_folder, f"{base_name}_Compared.xlsx")
        df.to_excel(output_file, index=False)
        auto_adjust_columns(output_file)
        print(f"✅ Saved: {output_file} ({found_count}/{total_rows} found)")

        summary_rows.append({
            "Master File": base_name,
            "Total Rows": total_rows,
            "EPCs Found": found_count,
            "EPCs Not Found": total_rows - found_count,
            "% Found": f"{percent_found:.2f}%"
        })

    except Exception as e:
        print(f"❌ Error processing {master_file}: {e}")

def save_summary_file(summary_rows, totals, output_folder):
    """Save summary Excel file"""
    if totals["total_rows"] > 0:
        overall_percent = (totals["found_rows"] / totals["total_rows"]) * 100
    else:
        overall_percent = 0.0

    # Add grand total row
    summary_rows.append({
        "Master File": "ALL FILES",
        "Total Rows": totals["total_rows"],
        "EPCs Found": totals["found_rows"],
        "EPCs Not Found": totals["total_rows"] - totals["found_rows"],
        "% Found": f"{overall_percent:.2f}%"
    })

    summary_df = pd.DataFrame(summary_rows)
    summary_file = os.path.join(output_folder, "Master_Database_Comparison_Summary.xlsx")
    summary_df.to_excel(summary_file, index=False)
    auto_adjust_columns(summary_file)
    print(f"📄 Summary file saved: {summary_file}")

def compare_epcs(merged_files, master_files):
    # Load merged EPCs with metadata
    merged_epcs = load_merged_epcs(merged_files)
    print(f"📦 Total unique EPCs from merged files: {len(merged_epcs)}")

    if not merged_epcs:
        print("❌ No EPCs loaded from merged files.")
        return

    # Ask user for mode: Merge or Per-file
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    merge_choice = messagebox.askyesno(
        "Processing Mode",
        "Do you want to merge all master files into one (Yes) or process them separately (No)?"
    )

    # Prepare output folder
    os.makedirs("comparison_results", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_folder = os.path.join("comparison_results", f"Master_Comparison_{timestamp}")
    os.makedirs(output_folder, exist_ok=True)

    summary_rows = []
    totals = {"total_rows": 0, "found_rows": 0}

    if merge_choice:
        # Merge all master files into one DataFrame
        dfs = []
        for file in master_files:
            try:
                if file.lower().endswith(".csv"):
                    df = pd.read_csv(file)
                else:
                    df = pd.read_excel(file)

                if "EPC" not in df.columns:
                    print(f"⚠️ Skipping {file} — no EPC column found.")
                    continue

                df["EPC"] = df["EPC"].astype(str).str.strip()
                df["Found"] = df["EPC"].apply(lambda x: "✅" if x in merged_epcs else "")
                df["Location Found"] = df["EPC"].apply(lambda x: merged_epcs[x]["Location"] if x in merged_epcs else "")
                df["Reader Used"] = df["EPC"].apply(lambda x: merged_epcs[x]["Reader"] if x in merged_epcs else "")

                found_in_file = df["Found"].eq("✅").sum()
                totals["total_rows"] += len(df)
                totals["found_rows"] += found_in_file

                dfs.append(df)
                print(f"🔍 Compared {file}: {found_in_file}/{len(df)} found")

            except Exception as e:
                print(f"❌ Error reading {file}: {e}")

        if dfs:
            combined_master = pd.concat(dfs, ignore_index=True)

            # Save as Excel or CSV based on row limit
            if len(combined_master) > EXCEL_MAX_ROWS:
                output_file = os.path.join(output_folder, "Master_Comparison_Combined.csv")
                combined_master.to_csv(output_file, index=False)
                print(f"📁 Dataset too large for Excel. Saved as CSV: {output_file}")
            else:
                output_file = os.path.join(output_folder, "Master_Comparison_Combined.xlsx")
                combined_master.to_excel(output_file, index=False)
                auto_adjust_columns(output_file)
                print(f"📁 Output saved as Excel: {output_file}")

            percent_found = (totals["found_rows"] / totals["total_rows"]) * 100 if totals["total_rows"] else 0
            summary_rows.append({
                "Master File": "All Merged",
                "Total Rows": totals["total_rows"],
                "EPCs Found": totals["found_rows"],
                "EPCs Not Found": totals["total_rows"] - totals["found_rows"],
                "% Found": f"{percent_found:.2f}%"
            })
        else:
            print("❌ No valid master files to merge.")

    else:
        # Process each master file separately
        for file in master_files:
            process_master_file(file, merged_epcs, output_folder, summary_rows, totals)

    # Save summary
    save_summary_file(summary_rows, totals, output_folder)

    # Popup summary
    messagebox.showinfo(
        "Comparison Complete",
        f"Comparison complete!\nResults saved to:\n{output_folder}"
    )

if __name__ == "__main__":
    print("📂 EPC Master Comparison Tool v9")

    # Select merged files
    merged_files = select_files_and_folders("Final Merged EPC Files")
    if not merged_files:
        print("❌ No merged files selected.")
        exit()

    # Select master database files
    master_files = select_files_and_folders("Master Database Files")
    if not master_files:
        print("❌ No master database files selected.")
        exit()

    compare_epcs(merged_files, master_files)
